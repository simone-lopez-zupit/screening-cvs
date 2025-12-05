"""
Estrae informazioni strutturate dai PDF dei CV, salva i risultati in Excel
e organizza zip dei CV accettati/rifiutati usando GPT-4o per il parsing.
"""

import argparse
import base64
import hashlib
import json
import shutil
import time
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openai import OpenAI
from dotenv import load_dotenv


SYSTEM_PROMPT = (
    "Sei un assistente che esegue OCR e parsing di CV. "
    "Analizza il PDF allegato e restituisci SOLO un oggetto JSON conforme allo schema richiesto. "
    "Non aggiungere testo fuori dal JSON."
)

USER_PROMPT = (
    "estrarre informazioni strutturate\n"
    "valutare 4 condizioni con spiegazione'\n"
    "determinare se il candidato e' ACCETTATO o RIFIUTATO\n"
    "INFORMAZIONI DA ESTRARRE PER OGNI CV\n"
    "- Nome completo della persona\n"
    "- Posizione lavorativa attuale\n"
    "- Luogo di residenza / citta'\n"
    "- Email\n"
    "- Telefono\n"
    "- Link LinkedIn (se presente)\n"
    "- Link GitHub (se presente)\n"
    "- Progetti personali extra-lavorativi citati nel CV (se presenti)\n"
    "- Esperienze lavorative o formative in settori diversi dallo sviluppo software (se presenti)\n"
    "Se un dato non e' ricavabile lascia la stringa vuota.\n"
    "\n"
    "2. VALUTAZIONE DELLE QUATTRO CONDIZIONI\n"
    "Per ciascuna condizione restituisci value, spiegazione:\n"
    "value: TRUE se verificata, FALSE se non verificata, NULL se non determinabile.\n"
    "spiegazione: descrivi esattamente come sei arrivato alla valutazione (testo trovato, "
    "assunzioni, deduzioni ecc.).\n"
    "\n"
    "3. DEFINIZIONE DELLE CONDIZIONI\n"
    "- ETA: TRUE se la persona NON ha piu' di 45 anni.\n"
    "- BOOLEAN: TRUE se NON ha frequentato bootcamp (Boolean Careers, Epicode, 42 School, "
    "Start2Impact, Develhope, Aulab, Ironhack, Le Wagon, ecc.).\n"
    "- ACCENTURE: TRUE se NON ha piu' di 5 anni complessivi in societa' di consulenza IT "
    "(Almaviva, Reply, Accenture, Deloitte, KPMG, Engineering, DXC, NTT Data, Capgemini, Everis, Sogeti, ecc.).\n"
    "- ITALIANO: TRUE se parla italiano come madrelingua o livello molto alto (C1/C2).\n"
    "\n"
    "4. CRITERIO DI ACCETTAZIONE\n"
    "Decisione ACCETTATO se nessuna condizione e' FALSE (possono essere TRUE o NULL). "
    "Decisione RIFIUTATO se almeno una condizione e' FALSE.\n"
    "\n"
    "Restituisci un oggetto JSON con questa struttura:\n"
    "{\n"
    '  "full_name": "",\n'
    '  "current_position": "",\n'
    '  "location": "",\n'
    '  "email": "",\n'
    '  "phone": "",\n'
    '  "linkedin": "",\n'
    '  "github": "",\n'
    '  "personal_projects": "",\n'
    '  "extra_tech": "",\n'
    '  "conditions": {\n'
    '    "eta": {"value": "TRUE|FALSE|NULL", "explanation": ""},\n'
    '    "boolean": {"value": "TRUE|FALSE|NULL", "explanation": ""},\n'
    '    "accenture": {"value": "TRUE|FALSE|NULL", "explanation": ""},\n'
    '    "italiano": {"value": "TRUE|FALSE|NULL", "explanation": ""}\n'
    "  },\n"
    '  "decision": "ACCETTATO|RIFIUTATO"\n'
    "}\n"
    "Lascia eventuali stringhe vuote per i dati mancanti e NON inventare informazioni."
)


def call_model_with_pdf_file(client: OpenAI, pdf_path: Path, model: str) -> Dict[str, str]:
    """
    Legge il PDF, lo codifica in base64 e lo passa al modello
    tramite Chat Completions API usando il tipo di contenuto 'file'.
    Il modello é forzato a rispondere in JSON tramite response_format.
    """
    # sono riuscito a mandarlo solo in base64 vabbu
    with pdf_path.open("rb") as f:
        pdf_bytes = f.read()
    base64_pdf = base64.b64encode(pdf_bytes).decode("utf-8")

    messages = [
        {
            "role": "system",
            "content": SYSTEM_PROMPT,
        },
        {
            "role": "user",
            "content": [
                {
                    "type": "file",
                    "file": {
                        "filename": pdf_path.name,
                        "file_data": f"data:application/pdf;base64,{base64_pdf}",
                    },
                },
                {
                    "type": "text",
                    "text": USER_PROMPT,
                },
            ],
        },
    ]

    completion = client.chat.completions.create(
        model=model,
        temperature=0,
        response_format={"type": "json_object"},
        messages=messages,
    )

    content = completion.choices[0].message.content

    try:
        return json.loads(content)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"Risposta non JSON dal modello: {e}; content={content!r}") from e


CONDITION_KEYS = ["eta", "boolean", "accenture", "italiano"]


def hash_file(path: Path, chunk_size: int = 1_048_576) -> str:
    """Calcola SHA-256 del file leggendo a chunk."""
    digest = hashlib.sha256()
    with path.open("rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def find_duplicates(input_dir: Path) -> Dict[str, List[Path]]:
    """Restituisce una mappa hash -> lista di file con lo stesso hash (solo duplicati)."""
    hashes: Dict[str, List[Path]] = {}
    for pdf_path in sorted(input_dir.iterdir()):
        if pdf_path.suffix.lower() != ".pdf":
            continue
        file_hash = hash_file(pdf_path)
        hashes.setdefault(file_hash, []).append(pdf_path)
    return {h: paths for h, paths in hashes.items() if len(paths) > 1}


def _unique_destination(dest_dir: Path, original_name: str) -> Path:
    """Trova un nome unico nella cartella di destinazione."""
    dest = dest_dir / original_name
    if not dest.exists():
        return dest
    stem = dest.stem
    suffix = dest.suffix
    counter = 1
    while True:
        candidate = dest_dir / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def move_duplicates(duplicates: Dict[str, List[Path]], target_dir: Path) -> None:
    """Sposta tutti i duplicati (tranne il primo di ogni gruppo) nella cartella target."""
    target_dir.mkdir(parents=True, exist_ok=True)
    for paths in duplicates.values():
        # Mantieni il primo file in origine, sposta gli altri
        for pdf_path in paths[1:]:
            destination = _unique_destination(target_dir, pdf_path.name)
            shutil.move(str(pdf_path), destination)


def sanitize_fields(raw: Dict[str, Any]) -> Dict[str, str]:
    """Normalizza i campi attesi, condizioni e decisione."""
    fields = ["full_name", "current_position", "location", "email", "phone", "linkedin", "github", "personal_projects", "extra_tech"]
    cleaned: Dict[str, str] = {field: (str(raw.get(field, "")) or "").strip() for field in fields}

    conditions_block = raw.get("conditions") or {}
    condition_values = []
    for key in CONDITION_KEYS:
        condition_data = conditions_block.get(key) or {}
        value = (str(condition_data.get("value", "") or "")).strip().upper()
        explanation = (str(condition_data.get("explanation", "") or "")).strip()

        cleaned[f"{key}_value"] = value
        cleaned[f"{key}_explanation"] = explanation

        if value:
            condition_values.append(value)

    raw_decision = (str(raw.get("decision") or "")).strip().upper()
    computed_decision = ""
    if condition_values:
        computed_decision = "RIFIUTATO" if "FALSE" in condition_values else "ACCETTATO"

    cleaned["decision"] = raw_decision or computed_decision
    return cleaned


OUTPUT_FIELDS = [
    "file_name",
    "full_name",
    "current_position",
    "location",
    "email",
    "phone",
    "linkedin",
    "github",
    "personal_projects",
    "extra_tech",
    "eta_value",
    "eta_explanation",
    "boolean_value",
    "boolean_explanation",
    "accenture_value",
    "accenture_explanation",
    "italiano_value",
    "italiano_explanation",
    "decision",
    "note",
]


def write_rows_to_excel(rows: List[Dict[str, str]], output_path: Path, headers: List[str]) -> None:
    """Salva le righe su un file Excel applicando il colore sulla decisione."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CV"
    ws.append(headers)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    print(f"Excel rows number: {len(rows)}\n")
    for row in rows:
        ws.append([row.get(field, "") for field in headers])
        current_row = ws.max_row
        decision_value = (row.get("decision") or "").upper()
        fill = green_fill if decision_value == "ACCETTATO" else red_fill if decision_value == "RIFIUTATO" else None
        if fill:
            for cell in ws[current_row]:
                cell.fill = fill

    wb.save(output_path)


def create_zip(zip_path: Path, files: List[Path]) -> None:
    """Crea uno zip contenente i file indicati; ignora i file mancanti."""
    with zipfile.ZipFile(zip_path, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file_path in files:
            if file_path.exists():
                zf.write(file_path, arcname=file_path.name)


def process_directory(
    input_dir: Path,
    model: str,
    pause: float,
    limit: Optional[int],
) -> List[Dict[str, str]]:
    files = sorted(p for p in input_dir.iterdir() if p.suffix.lower() == ".pdf")
    if limit is not None:
        files = files[:limit]

    client = OpenAI()
    rows: List[Dict[str, str]] = []
    for idx, pdf_path in enumerate(files, start=1):
        print(f"[{idx}/{len(files)}] Lavoro su: {pdf_path.name}")
        note = ""
        data: Dict[str, str] = {}
        try:
            raw = call_model_with_pdf_file(client, pdf_path, model)
            data = sanitize_fields(raw)
        except Exception as exc:  # noqa: BLE001
            note = f"errore: {exc}"
            data = sanitize_fields({})

        row = {"file_name": pdf_path.name, **data, "note": note}
        rows.append(row)

        if pause > 0 and idx < len(files):
            time.sleep(pause)

    return rows


def main() -> None:
    load_dotenv()
    parser = argparse.ArgumentParser(
        description="Estrae dati dai CV PDF e li salva in Excel/zip usando OpenAI GPT-4o."
    )
    parser.add_argument(
        "--input-dir",
        default="cvs",
        help="Cartella che contiene i PDF (default: cvs)",
    )
    parser.add_argument(
        "--duplicates-dir",
        default="cvs_duplicati",
        help="Cartella dove spostare i duplicati prima di processare (default: cvs_duplicati)",
    )
    parser.add_argument(
        "--model",
        default="gpt-4o",
        help="Modello OpenAI da usare (default: gpt-4o)",
    )
    parser.add_argument(
        "--pause",
        type=float,
        default=0.0,
        help="Secondi di attesa tra le chiamate API (default: 0)",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Numero massimo di PDF da processare (default: tutti)",
    )
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    if not input_dir.is_dir():
        raise SystemExit(f"Cartella di input non trovata: {input_dir}")

    duplicates_dir = Path(args.duplicates_dir)
    duplicates = find_duplicates(input_dir)
    if duplicates:
        print("Duplicati trovati (hash -> file):")
        for file_hash, paths in duplicates.items():
            print(f"\n{file_hash}:")
            for p in paths:
                print(f"  - {p.name}")
        move_duplicates(duplicates, duplicates_dir)
        print(f"\nDuplicati spostati in: {duplicates_dir}")
    else:
        print("Nessun duplicato trovato.")

    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    rows = process_directory(
        input_dir=input_dir,
        model=args.model,
        pause=args.pause,
        limit=args.limit,
    )

    excel_path = Path(f"cv_{timestamp_str}.xlsx")
    write_rows_to_excel(rows, output_path=excel_path, headers=OUTPUT_FIELDS)
    print(f"Excel salvato in: {excel_path}")

    accepted_files: List[Path] = []
    rejected_files: List[Path] = []
    for row in rows:
        decision = (row.get("decision") or "").upper()
        pdf_path = input_dir / row.get("file_name", "")
        if decision == "ACCETTATO":
            accepted_files.append(pdf_path)
        elif decision == "RIFIUTATO":
            rejected_files.append(pdf_path)

    zip_accept_path = Path(f"cv_approvati_{timestamp_str}.zip")
    zip_reject_path = Path(f"cv_rifiutati_{timestamp_str}.zip")
    create_zip(zip_accept_path, accepted_files)
    create_zip(zip_reject_path, rejected_files)
    print(f"Zip ACCETTATI: {zip_accept_path}")
    print(f"Zip RIFIUTATI: {zip_reject_path}")


if __name__ == "__main__":
    main()