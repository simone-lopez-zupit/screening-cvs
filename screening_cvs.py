"""
Estrae informazioni strutturate dai PDF dei CV, salva i risultati in Excel
e organizza zip dei CV accettati/rifiutati usando GPT-4o per il parsing.
"""

import base64
import json
import os
import shutil
import time
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openai import OpenAI
from dotenv import load_dotenv

from services.file_utils import hash_file
from services.manatal_service import build_headers, get_candidate_info
from find_duplicate_cvs import find_duplicates_by_hash


load_dotenv()

# ── Configuration ─────────────────────────────────────────────────────
INPUT_DIR = "cvs_confronto"
DUPLICATES_DIR = "cvs_duplicati"
MODEL = "gpt-4o"
PAUSE = 0.0
LIMIT = None
# ──────────────────────────────────────────────────────────────────

SYSTEM_PROMPT = (
    "Sei un assistente specializzato nell’elaborazione OCR e nel parsing di CV in formato PDF.\n"
    "Il tuo compito è leggere, interpretare e strutturare le informazioni presenti nel CV.\n"
    "Non aggiungere testo al di fuori dell’oggetto JSON.\n"
    "Non inventare o derivare dati che non sono esplicitamente presenti.\n"
    "Se un’informazione non è chiaramente indicata, lascia "" (stringa vuota) o null (per valori numerici).\n"
    "----------\n"
    "## FORMATO DATA E PERMANENZA (OBBLIGATORIO)\n"
    "Converti SEMPRE i periodi in anni decimali:\n"
    "- '2019-presente', '2019-oggi' → 2019-2026 = 7.0 anni\n"  
    "- '2021-attualmente' → 2021-2026 = 5.0 anni\n"
    "- 'Gennaio 2020 - presente' → 2020-2026 = 6.2 anni\n"
    "- Periodo singolo → 1.0 anno\n"
    "- Somma periodi multipli per stessa azienda\n"
    "----------\n"
)

USER_PROMPT = (
    "Estrai informazioni strutturate dal CV.\n"
    "\n"
    "INFORMAZIONI DA ESTRARRE:\n"
    "- Nome completo della persona\n"
    "- Posizione lavorativa attuale\n"
    "- Luogo di residenza / città\n"
    "- Email\n"
    "- Telefono\n"
    "- Link LinkedIn (se presente)\n"
    "- Link GitHub (se presente)\n"
    "- Progetti personali extra-lavorativi citati nel CV (se presenti)\n"
    "- Esperienze lavorative o formative in settori diversi dallo sviluppo software (se presenti)\n"
    "- Se ha almeno 3 anni di esperienza fullstack nello sviluppo web\n"
    "- Anno di nascita (se presente o deducibile)\n"
    "- Lingua in cui è scritto il CV\n"
    "- Lingue conosciute con livello (madrelingua, A1, A2, B1, B2, C1, C2)\n"
    "- Percorso di formazione: lista di istituti/scuole/bootcamp con tipo (università, bootcamp, corso online, ecc.)\n"
    "- Esperienze lavorative: lista di aziende con anni di permanenza\n"
    "Se un dato non é ricavabile lascia la stringa vuota o null.\n"
    "----------\n"
    "Restituisci un oggetto JSON con questa struttura:\n"
    "{\n"
    '  "full_name": "",\n'
    '  "current_position": "",\n'
    '  "location": "",\n'
    '  "email": "",\n'
    '  "phone": "",\n'
    '  "linkedin": "",\n'
    '  "github": "",\n'
    '  "personal_projects": "",\n'
    '  "extra_tech": "",\n'
    '  "3y_exp_web": "",\n'
    '  "birth_year": null,\n'
    '  "cv_language": "",\n'
    '  "languages": [{"language": "", "level": ""}],\n'
    '  "education": [{"institution": "", "type": ""}],\n'
    '  "work_experiences": [{"company": "", "years": 0.0}]\n'
    "}\n"
    "NON inventare informazioni. Estrai solo ciò che è presente nel CV."
)


def call_model_with_pdf_file(client: OpenAI, pdf_path: Path, model: str) -> Dict[str, str]:
    """
    Legge il PDF, lo codifica in base64 e lo passa al modello
    tramite Chat Completions API usando il tipo di contenuto 'file'.
    Il modello é forzato a rispondere in JSON tramite response_format.
    """
    # sono riuscito a mandarlo solo in base64 vabbu
    with pdf_path.open("rb") as f:
        pdf_bytes = f.read()
    base64_pdf = base64.b64encode(pdf_bytes).decode("utf-8")

    messages = [
        {
            "role": "system",
            "content": SYSTEM_PROMPT,
        },
        {
            "role": "user",
            "content": [
                {
                    "type": "file",
                    "file": {
                        "filename": pdf_path.name,
                        "file_data": f"data:application/pdf;base64,{base64_pdf}",
                    },
                },
                {
                    "type": "text",
                    "text": USER_PROMPT,
                },
            ],
        },
    ]

    completion = client.chat.completions.create(
        model=model,
        temperature=0,
        response_format={"type": "json_object"},
        messages=messages,
    )

    content = completion.choices[0].message.content

    try:
        return json.loads(content)
    except json.JSONDecodeError as e:
        raise RuntimeError(f"Risposta non JSON dal modello: {e}; content={content!r}") from e


CONDITION_KEYS = ["eta", "boolean", "accenture", "italiano"]

BOOTCAMP_NAMES = {
    "boolean careers", "boolean", "epicode", "42 school", "42",
    "start2impact", "develhope", "aulab", "ironhack", "le wagon", "digichamps",
}

CONSULTING_FIRMS = {
    "almaviva", "almawave", "reply", "accenture", "deloitte", "kpmg",
    "engineering", "dxc", "ntt data", "capgemini", "everis", "sogeti",
}


def evaluate_eta(raw: Dict[str, Any]) -> tuple:
    """Età: FALSE se >= 45 anni, TRUE se < 45, NULL se non determinabile."""
    birth_year = raw.get("birth_year")
    if birth_year is None:
        return "NULL", "Anno di nascita non presente nel CV."
    try:
        birth_year = int(birth_year)
    except (ValueError, TypeError):
        return "NULL", f"Anno di nascita non valido: {birth_year}"
    current_year = datetime.now().year
    age = current_year - birth_year
    if age >= 45:
        return "FALSE", f"Nato nel {birth_year}, età stimata {age} (>= 45)."
    return "TRUE", f"Nato nel {birth_year}, età stimata {age} (< 45)."


def evaluate_boolean(raw: Dict[str, Any]) -> tuple:
    """Boolean: FALSE se ha frequentato un bootcamp noto, TRUE altrimenti."""
    education = raw.get("education") or []
    for entry in education:
        institution = (str(entry.get("institution", "")) or "").strip().lower()
        for bootcamp in BOOTCAMP_NAMES:
            if bootcamp in institution:
                return "FALSE", f"Ha frequentato il bootcamp: {entry.get('institution')}."
    return "TRUE", "Nessun bootcamp noto trovato nel percorso formativo."


def evaluate_accenture(raw: Dict[str, Any]) -> tuple:
    """Accenture: FALSE se > 5 anni complessivi in società di consulenza IT."""
    work_experiences = raw.get("work_experiences") or []
    total_years = 0.0
    matched_companies = []
    for entry in work_experiences:
        company = (str(entry.get("company", "")) or "").strip().lower()
        for firm in CONSULTING_FIRMS:
            if firm in company:
                years = 0.0
                try:
                    years = float(entry.get("years", 0) or 0)
                except (ValueError, TypeError):
                    pass
                total_years += years
                matched_companies.append(f"{entry.get('company')} ({years}a)")
                break
    if total_years > 5:
        return "FALSE", f"Totale {total_years} anni in consulenza IT: {', '.join(matched_companies)}."
    if matched_companies:
        return "TRUE", f"Totale {total_years} anni in consulenza IT (<= 5): {', '.join(matched_companies)}."
    return "TRUE", "Nessuna esperienza in società di consulenza IT note."


def evaluate_italiano(raw: Dict[str, Any]) -> tuple:
    """Italiano: TRUE se parla italiano madrelingua/C1/C2 o CV scritto in italiano."""
    cv_language = (str(raw.get("cv_language", "")) or "").strip().lower()
    if cv_language in ("italiano", "italian", "it"):
        return "TRUE", "Il CV è scritto in italiano."

    languages = raw.get("languages") or []
    for entry in languages:
        lang = (str(entry.get("language", "")) or "").strip().lower()
        level = (str(entry.get("level", "")) or "").strip().lower()
        if lang in ("italiano", "italian", "it"):
            if any(kw in level for kw in ("madrelingua", "nativo", "native", "c1", "c2")):
                return "TRUE", f"Italiano dichiarato a livello: {entry.get('level')}."
            return "FALSE", f"Italiano dichiarato a livello: {entry.get('level')} (inferiore a C1)."

    return "FALSE", "Italiano non menzionato e CV non in italiano."



def find_duplicates(input_dir: Path) -> Dict[str, List[Path]]:
    """Restituisce una mappa hash -> lista di file con lo stesso hash (solo duplicati)."""
    hashes: Dict[str, List[Path]] = {}
    for pdf_path in sorted(input_dir.iterdir()):
        if pdf_path.suffix.lower() != ".pdf":
            continue
        file_hash = hash_file(pdf_path)
        hashes.setdefault(file_hash, []).append(pdf_path)
    return {h: paths for h, paths in hashes.items() if len(paths) > 1}


def _unique_destination(dest_dir: Path, original_name: str) -> Path:
    """Trova un nome unico nella cartella di destinazione."""
    dest = dest_dir / original_name
    if not dest.exists():
        return dest
    stem = dest.stem
    suffix = dest.suffix
    counter = 1
    while True:
        candidate = dest_dir / f"{stem}_{counter}{suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def move_duplicates(duplicates: Dict[str, List[Path]], target_dir: Path) -> None:
    """Sposta tutti i duplicati (tranne il primo di ogni gruppo) nella cartella target."""
    target_dir.mkdir(parents=True, exist_ok=True)
    for paths in duplicates.values():
        # Mantieni il primo file in origine, sposta gli altri
        for pdf_path in paths[1:]:
            destination = _unique_destination(target_dir, pdf_path.name)
            shutil.move(str(pdf_path), destination)


EVALUATE_FUNCTIONS = {
    "eta": evaluate_eta,
    "boolean": evaluate_boolean,
    "accenture": evaluate_accenture,
    "italiano": evaluate_italiano,
}


def sanitize_fields(raw: Dict[str, Any]) -> Dict[str, str]:
    """Normalizza i campi attesi, valuta condizioni in Python e determina decisione."""
    fields = ["full_name", "current_position", "location", "email", "phone", "linkedin", "github", "personal_projects", "extra_tech", "3y_exp_web"]
    cleaned: Dict[str, str] = {field: (str(raw.get(field, "")) or "").strip() for field in fields}

    condition_values = []
    for key in CONDITION_KEYS:
        evaluate_fn = EVALUATE_FUNCTIONS[key]
        value, explanation = evaluate_fn(raw)
        cleaned[f"{key}_value"] = value
        cleaned[f"{key}_explanation"] = explanation
        condition_values.append(value)

    cleaned["decision"] = "RIFIUTATO" if "FALSE" in condition_values else "ACCETTATO"
    return cleaned


OUTPUT_FIELDS = [
    "file_name",
    "full_name",
    "current_position",
    "location",
    "email",
    "phone",
    "linkedin",
    "github",
    "personal_projects",
    "extra_tech",
    "3y_exp_web",
    "eta_value",
    "eta_explanation",
    "boolean_value",
    "boolean_explanation",
    "accenture_value",
    "accenture_explanation",
    "italiano_value",
    "italiano_explanation",
    "decision",
    "manatal_link",
    "manatal_job",
    "manatal_stage",
    "manatal_is_dropped",
    "manatal_drop_date",
    "is_duplicate",
    "note",
]


def write_rows_to_excel(rows: List[Dict[str, str]], output_path: Path, headers: List[str]) -> None:
    """Salva le righe su un file Excel applicando il colore sulla decisione."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CV"
    ws.append(headers)

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    print(f"Excel rows number: {len(rows)}\n")
    for row in rows:
        ws.append([row.get(field, "") for field in headers])
        current_row = ws.max_row
        decision_value = (row.get("decision") or "").upper()
        fill = green_fill if decision_value == "ACCETTATO" else red_fill if decision_value == "RIFIUTATO" else None
        if fill:
            for cell in ws[current_row]:
                cell.fill = fill

    wb.save(output_path)


def create_zip(zip_path: Path, files: List[Path]) -> None:
    """Crea uno zip contenente i file indicati; ignora i file mancanti."""
    with zipfile.ZipFile(zip_path, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for file_path in files:
            if file_path.exists():
                zf.write(file_path, arcname=file_path.name)


def _build_processed_filenames(processed_dir: Path) -> set:
    """Collect all PDF filenames from cvs_processed/ for duplicate detection."""
    names = set()
    if processed_dir.is_dir():
        for pdf in processed_dir.rglob("*.pdf"):
            names.add(pdf.name)
    return names


DUPLICATE_CUTOFF = "2026-01-01"

ROLE_KEYWORDS = {
    "TL": "TL",
    "Jun Dev": "Jun Dev",
    "Jun Mid": "Mid Dev",
    "Jun Sen": "Sen Dev",
    "DEV Sen": "Sen Dev",
    "DEV Mid": "Mid Dev",
    "DEV Jun": "Jun Dev",
}


def _detect_role(subfolder: Path) -> str:
    """Detect the role from the first CV filename in the subfolder."""
    first_pdf = next((p for p in sorted(subfolder.iterdir()) if p.suffix.lower() == ".pdf"), None)
    if not first_pdf:
        return ""
    name = first_pdf.name
    # Filenames follow "CV - <ROLE> - ..." pattern
    for keyword, role in ROLE_KEYWORDS.items():
        if f"- {keyword} -" in name or f"- {keyword} " in name:
            return role
    return ""


def process_directory(
    headers: Dict[str, str],
    input_dir: Path,
    model: str,
    pause: float,
    limit: Optional[int],
    processed_filenames: set = None,
) -> List[Dict[str, str]]:
    files = sorted(p for p in input_dir.iterdir() if p.suffix.lower() == ".pdf")
    if limit is not None:
        files = files[:limit]

    client = OpenAI()
    rows: List[Dict[str, str]] = []
    for idx, pdf_path in enumerate(files, start=1):
        print(f"[{idx}/{len(files)}] Lavoro su: {pdf_path.name}")
        note = ""

        raw = {}
        try:
            raw = call_model_with_pdf_file(client, pdf_path, model)
            data = sanitize_fields(raw)
        except Exception as exc:  # noqa: BLE001
            note = f"errore: {exc}"
            data = sanitize_fields({})

        cand_email = raw.get("email")
        created_at = None
        if cand_email:
            manatal_link, match_details, created_at = get_candidate_info(headers, cand_email)
        else:
            manatal_link, match_details = "", []

        manatal_jobs = "\n".join(m["job"] for m in match_details) if match_details else ""
        manatal_stages = "\n".join(m["stage"] for m in match_details) if match_details else ""
        manatal_dropped = "\n".join(str(m["is_dropped"]) for m in match_details) if match_details else ""
        manatal_drop_dates = "\n".join(m["drop_date"] for m in match_details) if match_details else ""

        # ── Duplicate detection ──────────────────────────────────────
        is_duplicate = False
        if created_at and created_at[:10] >= DUPLICATE_CUTOFF:
            is_duplicate = True
        if processed_filenames and pdf_path.name in processed_filenames:
            is_duplicate = True

        row = {
            "file_name": pdf_path.name,
            **data,
            "manatal_link": manatal_link,
            "manatal_job": manatal_jobs,
            "manatal_stage": manatal_stages,
            "manatal_is_dropped": manatal_dropped,
            "manatal_drop_date": manatal_drop_dates,
            "is_duplicate": is_duplicate,
            "note": note,
        }
        rows.append(row)

        if pause > 0 and idx < len(files):
            time.sleep(pause)

    return rows


def main() -> None:
    input_dir = Path(INPUT_DIR)

    if not input_dir.is_dir():
        raise SystemExit(f"Cartella di input non trovata: {input_dir}")

    # ── Cross-folder duplicate detection (grouped by role) ─────────────
    subfolders = sorted(p for p in input_dir.iterdir() if p.is_dir() and p.name != "cvs_processed")
    role_groups: Dict[str, List[Path]] = {}
    for sf in subfolders:
        role = _detect_role(sf)
        role_groups.setdefault(role, []).append(sf)

    print("=== Controllo duplicati tra sottocartelle (per ruolo) ===\n")
    any_dups = False
    for role, folders in role_groups.items():
        if len(folders) < 2:
            continue
        hash_dups = find_duplicates_by_hash(input_dir, folders=folders)
        if hash_dups:
            any_dups = True
            label = role or "sconosciuto"
            print(f"{len(hash_dups)} CV identici ({label}) in più cartelle:\n")
            for paths in hash_dups.values():
                print(f"  {paths[0].name}")
                for p in paths:
                    print(f"    - {p.parent.name}")
                print()

    if not any_dups:
        print("Nessun CV duplicato trovato tra le sottocartelle.\n")
    if not subfolders:
        raise SystemExit(f"Nessuna sottocartella trovata in: {input_dir}")

    headers = build_headers()
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    processed_filenames = _build_processed_filenames(input_dir / "cvs_processed")

    for subfolder in subfolders:
        role = _detect_role(subfolder)
        role_prefix = f"{role}_" if role else ""
        print(f"\n{'='*60}")
        print(f"=== Screening: {role_prefix}{subfolder.name} ===")
        print(f"{'='*60}\n")

        # Dedup within subfolder
        duplicates_dir = subfolder / DUPLICATES_DIR
        duplicates = find_duplicates(subfolder)
        if duplicates:
            print("Duplicati interni trovati (hash -> file):")
            for file_hash, paths in duplicates.items():
                print(f"\n{file_hash}:")
                for p in paths:
                    print(f"  - {p.name}")
            move_duplicates(duplicates, duplicates_dir)
            print(f"\nDuplicati spostati in: {duplicates_dir}")
        else:
            print("Nessun duplicato interno trovato.")

        rows = process_directory(
            headers=headers,
            input_dir=subfolder,
            model=MODEL,
            pause=PAUSE,
            limit=LIMIT,
            processed_filenames=processed_filenames,
        )

        output_dir = Path(f"output_{role_prefix}{subfolder.name}_{timestamp_str}")
        output_dir.mkdir(exist_ok=True)

        excel_path = output_dir / f"cv_{role_prefix}{subfolder.name}_{timestamp_str}.xlsx"
        write_rows_to_excel(rows, output_path=excel_path, headers=OUTPUT_FIELDS)
        print(f"Excel salvato in: {excel_path}")

        accepted_files: List[Path] = []
        rejected_files: List[Path] = []
        duplicate_files: List[Path] = []

        for row in rows:
            pdf_path = subfolder / row.get("file_name", "")
            if row.get("is_duplicate"):
                duplicate_files.append(pdf_path)
                continue
            decision = (row.get("decision") or "").upper()
            if decision == "ACCETTATO":
                accepted_files.append(pdf_path)
            elif decision == "RIFIUTATO":
                rejected_files.append(pdf_path)

        # Move duplicate CVs to dedicated folder
        if duplicate_files:
            dup_dir = output_dir / "cv_duplicati"
            dup_dir.mkdir(exist_ok=True)
            for dup_path in duplicate_files:
                if dup_path.exists():
                    shutil.move(str(dup_path), str(dup_dir / dup_path.name))
            print(f"Duplicati spostati in: {dup_dir} ({len(duplicate_files)} file)")

        zip_accept_path = output_dir / f"cv_approvati_{role_prefix}{subfolder.name}_{timestamp_str}.zip"
        zip_reject_path = output_dir / f"cv_rifiutati_{role_prefix}{subfolder.name}_{timestamp_str}.zip"
        create_zip(zip_accept_path, accepted_files)
        create_zip(zip_reject_path, rejected_files)
        print(f"Zip ACCETTATI: {zip_accept_path}")
        print(f"Zip RIFIUTATI: {zip_reject_path}")
        print(f"\nOutput in: {output_dir}")

        # Move processed subfolder to cvs_processed
        processed_dest = input_dir / "cvs_processed" / subfolder.name
        (input_dir / "cvs_processed").mkdir(parents=True, exist_ok=True)
        shutil.move(str(subfolder), str(processed_dest))
        print(f"Cartella spostata in: {processed_dest}")


if __name__ == "__main__":
    main()
